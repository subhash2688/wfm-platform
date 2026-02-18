"""Export prospects back to Excel format."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from models.database import db_session
from models.prospect import Prospect
from config import EXCEL_PATH
import os


def export_to_excel(output_path=None):
    """Export all prospects to a new Excel file matching original format."""
    if output_path is None:
        base = os.path.dirname(EXCEL_PATH)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(base, f'Corporate_Prospect_Tracker_export_{timestamp}.xlsx')

    wb = openpyxl.Workbook()

    # --- Prospect Pipeline sheet ---
    ws = wb.active
    ws.title = 'Prospect Pipeline'

    headers = [
        'Company Name', 'Industry', 'HQ City', 'Nearest Campus',
        'Giving Channel', 'Focus Areas', 'Contact Name', 'Contact Title',
        'Contact Email', 'Contact LinkedIn', 'Alignment (1-5)', 'Proximity (1-5)',
        'Capacity (1-5)', 'Total Score', 'Pipeline Stage', 'Last Action',
        'Last Action Date', 'Next Step', 'Next Step Date', 'Ask Amount',
        'Amount Received', 'Notes',
    ]

    # Header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2D5016', end_color='2D5016', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    prospects = db_session.query(Prospect).order_by(Prospect.total_score.desc()).all()
    for row_idx, p in enumerate(prospects, 2):
        vals = [
            p.company_name, p.industry, p.hq_city, p.nearest_campus,
            p.giving_channel, p.focus_areas, p.contact_name, p.contact_title,
            p.contact_email, p.contact_linkedin, p.alignment_score, p.proximity_score,
            p.capacity_score, p.total_score, p.pipeline_stage, p.last_action,
            p.last_action_date, p.next_step, p.next_step_date,
            p.ask_amount or 0, p.amount_received or 0, p.notes,
        ]
        for col_idx, val in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

        # Add total score formula
        ws.cell(row=row_idx, column=14, value=f'=K{row_idx}+L{row_idx}+M{row_idx}')

    # Column widths
    widths = [25, 18, 16, 18, 20, 30, 20, 25, 28, 30, 12, 12, 12, 10, 15, 20, 14, 20, 14, 12, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Grant Deadlines sheet ---
    ws2 = wb.create_sheet('Grant Deadlines')
    dl_headers = ['Company / Foundation', 'Grant Program Name', 'Focus Area',
                  'Application Opens', 'Deadline', 'Typical Award Range',
                  'Application URL', 'Status', 'Notes']
    for col, header in enumerate(dl_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws2.freeze_panes = 'A2'

    # --- Dashboard sheet ---
    ws3 = wb.create_sheet('Dashboard')
    ws3.cell(row=1, column=1, value='WFM Corporate Fundraising Dashboard').font = Font(bold=True, size=14)
    ws3.cell(row=3, column=1, value='Pipeline Summary').font = Font(bold=True, size=12)

    metrics = [
        ('Total Prospects', len(prospects)),
        ('In Active Outreach', sum(1 for p in prospects if p.pipeline_stage in ['3-Outreach Sent', '4-Meeting Scheduled'])),
        ('Proposals Submitted', sum(1 for p in prospects if p.pipeline_stage == '5-Proposal Sent')),
        ('Funded', sum(1 for p in prospects if p.pipeline_stage == '7-Funded')),
        ('Total Ask Amount', sum(p.ask_amount or 0 for p in prospects)),
        ('Total Received', sum(p.amount_received or 0 for p in prospects)),
    ]
    for i, (label, val) in enumerate(metrics):
        ws3.cell(row=4 + i, column=1, value=label)
        ws3.cell(row=4 + i, column=2, value=val)

    wb.save(output_path)
    return output_path

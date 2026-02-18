"""Import prospects from Corporate_Prospect_Tracker.xlsx."""
import openpyxl
from datetime import datetime
from models.database import db_session
from models.prospect import Prospect
from models.activity_log import ActivityLog
from config import EXCEL_PATH


def import_from_excel(filepath=None):
    """Import prospects from Excel. Returns count of imported records."""
    filepath = filepath or EXCEL_PATH
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Prospect Pipeline']

    imported = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        company_name = ws.cell(row=row, column=1).value
        if not company_name:
            continue

        # Skip if already exists
        existing = db_session.query(Prospect).filter_by(company_name=company_name).first()
        if existing:
            skipped += 1
            continue

        prospect = Prospect(
            company_name=str(company_name).strip(),
            industry=_cell(ws, row, 2),
            hq_city=_cell(ws, row, 3),
            nearest_campus=_cell(ws, row, 4),
            giving_channel=_cell(ws, row, 5),
            focus_areas=_cell(ws, row, 6),
            contact_name=_cell(ws, row, 7),
            contact_title=_cell(ws, row, 8),
            contact_email=_cell(ws, row, 9),
            contact_linkedin=_cell(ws, row, 10),
            alignment_score=_int_cell(ws, row, 11),
            proximity_score=_int_cell(ws, row, 12),
            capacity_score=_int_cell(ws, row, 13),
            total_score=_int_cell(ws, row, 14),
            pipeline_stage=_cell(ws, row, 15) or '1-Research',
            last_action=_cell(ws, row, 16),
            last_action_date=_cell(ws, row, 17),
            next_step=_cell(ws, row, 18),
            next_step_date=_cell(ws, row, 19),
            ask_amount=_float_cell(ws, row, 20),
            amount_received=_float_cell(ws, row, 21),
            notes=_cell(ws, row, 22),
            research_status='imported',
        )
        prospect.recalculate_total()
        db_session.add(prospect)
        imported += 1

    if imported > 0:
        log = ActivityLog(
            action_type='import',
            description=f'Imported {imported} prospects from Excel ({skipped} skipped as duplicates)',
        )
        db_session.add(log)
        db_session.commit()

    wb.close()
    return imported, skipped


def _cell(ws, row, col):
    """Get cell value as string or None."""
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else None


def _int_cell(ws, row, col):
    """Get cell value as int or 0."""
    val = ws.cell(row=row, column=col).value
    try:
        return int(val)
    except (TypeError, ValueError):
        return 0


def _float_cell(ws, row, col):
    """Get cell value as float or 0."""
    val = ws.cell(row=row, column=col).value
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0
